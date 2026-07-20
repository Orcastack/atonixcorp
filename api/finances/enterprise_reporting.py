from collections import defaultdict
from datetime import timedelta, timezone as datetime_timezone
from decimal import Decimal
from io import BytesIO
import json
from zoneinfo import ZoneInfo

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.files.base import ContentFile
from django.db.models import Sum
from django.utils import timezone

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from equity.models import WorkspaceEquityProfile
from equity.scenario_services import get_scenario_overview, simulate_financing_scenario

from .platform_foundation import log_platform_audit_event

from .models import (
    AuditLog,
    AutomationArtifact,
    AutomationExecution,
    AutomationWorkflow,
    BankAccount,
    Budget,
    CashflowForecast,
    ComplianceDeadline,
    ComplianceDocument,
    Consolidation,
    RecurringJournalTemplate,
    RecurringTransaction,
    Scenario,
    TaxExposure,
    Transaction,
    Wallet,
)


DECIMAL_ZERO = Decimal('0.00')
DEFAULT_SCHEDULE_TIMEZONE = 'UTC'
DEFAULT_ARTIFACT_RETENTION_DAYS = 90
SUPPORTED_REPORT_ACTIONS = {
    'enterprise_reporting_pack',
    'enterprise_report_pack',
    'compliance_reporting_pack',
}


def _safe_decimal(value):
    if value in [None, '']:
        return DECIMAL_ZERO
    return Decimal(str(value))


def _serialize_money(value):
    return round(float(value or 0), 2)


def _month_key(value):
    return value.strftime('%Y-%m')


def _serialize_month_rows(month_map):
    ordered = []
    for month_key in sorted(month_map.keys()):
        row = month_map[month_key]
        projected = row['opening_cash'] + row['forecast_net']
        ordered.append({
            'month': month_key,
            'actual_inflow': _serialize_money(row['actual_inflow']),
            'actual_outflow': _serialize_money(row['actual_outflow']),
            'actual_net': _serialize_money(row['actual_net']),
            'forecast_inflow': _serialize_money(row['forecast_inflow']),
            'forecast_outflow': _serialize_money(row['forecast_outflow']),
            'forecast_net': _serialize_money(row['forecast_net']),
            'opening_cash': _serialize_money(row['opening_cash']),
            'projected_closing_cash': _serialize_money(projected),
        })
    return ordered


def _entity_statement_summary(entity, start, end):
    transactions = Transaction.objects.filter(entity=entity, date__gte=start, date__lte=end)
    revenue = _safe_decimal(transactions.filter(type='income').aggregate(total=Sum('amount'))['total'])
    expenses = _safe_decimal(transactions.filter(type='expense').aggregate(total=Sum('amount'))['total'])
    accrual_revenue = _safe_decimal(
        entity.accrual_entries.filter(
            accrual_type='revenue',
            accrual_date__gte=start,
            accrual_date__lte=end,
        ).aggregate(total=Sum('amount'))['total']
    )
    accrual_expenses = _safe_decimal(
        entity.accrual_entries.filter(
            accrual_type='expense',
            accrual_date__gte=start,
            accrual_date__lte=end,
        ).aggregate(total=Sum('amount'))['total']
    )
    depreciation = DECIMAL_ZERO
    for asset in entity.fixed_assets.filter(is_active=True):
        annual = _safe_decimal(asset.calculate_depreciation())
        months = max(1, ((end.year - start.year) * 12) + (end.month - start.month) + 1)
        depreciation += annual * Decimal(str(months / 12))

    net_income = revenue + accrual_revenue - expenses - accrual_expenses - depreciation
    cash_balance = _safe_decimal(entity.bank_accounts.aggregate(total=Sum('balance'))['total'])
    cash_balance += _safe_decimal(entity.wallets.aggregate(total=Sum('balance'))['total'])
    receivables = _safe_decimal(
        entity.invoices.exclude(status__in=['paid', 'cancelled']).aggregate(total=Sum('outstanding_amount'))['total']
    )
    liabilities = _safe_decimal(
        entity.bills.exclude(status__in=['paid', 'cancelled']).aggregate(total=Sum('outstanding_amount'))['total']
    )
    fixed_assets = DECIMAL_ZERO
    for asset in entity.fixed_assets.filter(is_active=True):
        fixed_assets += _safe_decimal(asset.cost) - _safe_decimal(asset.accumulated_depreciation)
    total_assets = cash_balance + receivables + fixed_assets
    equity = total_assets - liabilities

    return {
        'entity_id': entity.id,
        'entity_name': entity.name,
        'currency': entity.local_currency or entity.organization.primary_currency,
        'revenue': revenue + accrual_revenue,
        'expenses': expenses + accrual_expenses + depreciation,
        'net_income': net_income,
        'cash': cash_balance,
        'receivables': receivables,
        'fixed_assets': fixed_assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_assets': total_assets,
    }


def _build_cashflow_months(entities, start, end):
    from django.db.models.functions import TruncMonth

    month_map = {}
    current_month = start.replace(day=1)
    while current_month <= end.replace(day=1):
        month_key = _month_key(current_month)
        month_map[month_key] = {
            'actual_inflow': DECIMAL_ZERO,
            'actual_outflow': DECIMAL_ZERO,
            'actual_net': DECIMAL_ZERO,
            'forecast_inflow': DECIMAL_ZERO,
            'forecast_outflow': DECIMAL_ZERO,
            'forecast_net': DECIMAL_ZERO,
            'opening_cash': DECIMAL_ZERO,
        }
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)

    entity_ids = [entity.id for entity in entities]
    txn_rows = (
        Transaction.objects.filter(entity_id__in=entity_ids, date__gte=start, date__lte=end)
        .annotate(month=TruncMonth('date'))
        .values('month', 'type')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )
    for row in txn_rows:
        if not row.get('month'):
            continue
        month_key = _month_key(row['month'])
        payload = month_map.setdefault(month_key, {
            'actual_inflow': DECIMAL_ZERO,
            'actual_outflow': DECIMAL_ZERO,
            'actual_net': DECIMAL_ZERO,
            'forecast_inflow': DECIMAL_ZERO,
            'forecast_outflow': DECIMAL_ZERO,
            'forecast_net': DECIMAL_ZERO,
            'opening_cash': DECIMAL_ZERO,
        })
        total = _safe_decimal(row['total'])
        if row['type'] == 'income':
            payload['actual_inflow'] += total
            payload['actual_net'] += total
        elif row['type'] == 'expense':
            payload['actual_outflow'] += total
            payload['actual_net'] -= total

    forecast_rows = CashflowForecast.objects.filter(
        entity_id__in=entity_ids,
        month__gte=start.replace(day=1),
        month__lte=end.replace(day=1),
    )
    for forecast in forecast_rows:
        month_key = _month_key(forecast.month)
        payload = month_map.setdefault(month_key, {
            'actual_inflow': DECIMAL_ZERO,
            'actual_outflow': DECIMAL_ZERO,
            'actual_net': DECIMAL_ZERO,
            'forecast_inflow': DECIMAL_ZERO,
            'forecast_outflow': DECIMAL_ZERO,
            'forecast_net': DECIMAL_ZERO,
            'opening_cash': DECIMAL_ZERO,
        })
        amount = _safe_decimal(forecast.forecasted_amount)
        if amount >= 0:
            payload['forecast_inflow'] += amount
        else:
            payload['forecast_outflow'] += abs(amount)
        payload['forecast_net'] += amount

    opening_cash = _safe_decimal(BankAccount.objects.filter(entity_id__in=entity_ids).aggregate(total=Sum('balance'))['total'])
    opening_cash += _safe_decimal(Wallet.objects.filter(entity_id__in=entity_ids).aggregate(total=Sum('balance'))['total'])
    rolling_cash = opening_cash
    for month_key in sorted(month_map.keys()):
        month_map[month_key]['opening_cash'] = rolling_cash
        rolling_cash += month_map[month_key]['forecast_net'] or month_map[month_key]['actual_net']

    return _serialize_month_rows(month_map)


def _build_budget_section(entities, end):
    entity_ids = [entity.id for entity in entities]
    budgets = Budget.objects.filter(entity_id__in=entity_ids).select_related('entity')
    next_month_cutoff = (end.replace(day=1) + timedelta(days=32)).replace(day=1)
    forecasts = CashflowForecast.objects.filter(entity_id__in=entity_ids, month__gte=end.replace(day=1), month__lt=next_month_cutoff)

    rows = []
    total_limit = DECIMAL_ZERO
    total_spent = DECIMAL_ZERO
    total_forecast = DECIMAL_ZERO
    for budget in budgets:
        forecast_total = _safe_decimal(
            forecasts.filter(entity=budget.entity, category__iexact=budget.category).aggregate(total=Sum('forecasted_amount'))['total']
        )
        limit_amount = _safe_decimal(budget.limit)
        spent_amount = _safe_decimal(budget.spent)
        variance_amount = limit_amount - forecast_total
        total_limit += limit_amount
        total_spent += spent_amount
        total_forecast += forecast_total
        rows.append({
            'id': budget.id,
            'entity_id': budget.entity_id,
            'entity_name': budget.entity.name if budget.entity else 'Unscoped',
            'category': budget.category,
            'limit': _serialize_money(limit_amount),
            'spent': _serialize_money(spent_amount),
            'forecast': _serialize_money(forecast_total),
            'remaining': _serialize_money(limit_amount - spent_amount),
            'variance': _serialize_money(variance_amount),
            'utilization_percent': round(float((spent_amount / limit_amount) * 100), 2) if limit_amount else 0,
        })

    rows.sort(key=lambda row: abs(row['variance']), reverse=True)
    return {
        'summary': {
            'budget_limit': _serialize_money(total_limit),
            'actual_spend': _serialize_money(total_spent),
            'forecast_spend': _serialize_money(total_forecast),
            'variance': _serialize_money(total_limit - total_forecast),
        },
        'top_variances': rows[:8],
        'categories': rows,
    }


def _build_compliance_section(organization, entities, start, end):
    entity_ids = [entity.id for entity in entities]
    deadlines = ComplianceDeadline.objects.filter(organization=organization, entity_id__in=entity_ids).select_related('entity').order_by('deadline_date')
    exposures = TaxExposure.objects.filter(entity_id__in=entity_ids)
    documents = ComplianceDocument.objects.filter(entity_id__in=entity_ids)
    audit_events = AuditLog.objects.filter(organization=organization, created_at__date__gte=start, created_at__date__lte=end)

    status_counts = {
        'upcoming': deadlines.filter(status='upcoming').count(),
        'due_soon': deadlines.filter(status='due_soon').count(),
        'overdue': deadlines.filter(status='overdue').count(),
        'completed': deadlines.filter(status='completed').count(),
    }
    upcoming = [
        {
            'id': deadline.id,
            'entity_name': deadline.entity.name,
            'title': deadline.title,
            'deadline_date': deadline.deadline_date.isoformat(),
            'status': deadline.status,
            'deadline_type': deadline.deadline_type,
        }
        for deadline in deadlines[:10]
    ]

    return {
        'status_counts': status_counts,
        'upcoming_deadlines': upcoming,
        'tax_exposure': {
            'estimated': _serialize_money(exposures.aggregate(total=Sum('estimated_amount'))['total']),
            'actual': _serialize_money(exposures.aggregate(total=Sum('actual_amount'))['total']),
            'paid': _serialize_money(exposures.aggregate(total=Sum('paid_amount'))['total']),
        },
        'document_coverage': {
            'documents_on_file': documents.count(),
            'entities_with_documents': documents.values('entity_id').distinct().count(),
            'entity_coverage_percent': round((documents.values('entity_id').distinct().count() / len(entity_ids)) * 100, 2) if entity_ids else 0,
        },
        'automation': {
            'audit_events': audit_events.count(),
            'exports_logged': audit_events.filter(action='export').count(),
            'deadline_completion_rate': round((status_counts['completed'] / max(1, deadlines.count())) * 100, 2),
        },
    }


def _build_scenario_section(organization):
    scenarios = Scenario.objects.filter(financial_model__organization=organization).select_related('financial_model').order_by('-updated_at')
    top = []
    type_counts = defaultdict(int)
    for scenario in scenarios[:8]:
        type_counts[scenario.scenario_type] += 1
        top.append({
            'id': scenario.id,
            'name': scenario.name,
            'scenario_type': scenario.scenario_type,
            'financial_model_name': scenario.financial_model.name,
            'enterprise_value': _serialize_money(scenario.enterprise_value),
            'irr': round(float(scenario.irr or 0), 4),
            'probability': round(float(scenario.probability or 0), 2),
            'updated_at': scenario.updated_at.isoformat(),
        })
    return {
        'count': scenarios.count(),
        'by_type': dict(type_counts),
        'top_scenarios': top,
    }


def _default_equity_scenario_input(workspace, overview):
    share_class = workspace.equity_share_classes.order_by('-liquidation_seniority', 'name').first()
    if not share_class:
        return None

    defaults = overview.get('defaults', {})
    fully_diluted_shares = int((overview.get('cap_table') or {}).get('fully_diluted_shares') or 0)
    pre_money = _safe_decimal(defaults.get('latest_round_pre_money'))
    price_per_share = _safe_decimal(defaults.get('latest_round_price_per_share') or defaults.get('latest_valuation_price_per_share'))
    if pre_money <= 0 and price_per_share > 0 and fully_diluted_shares > 0:
        pre_money = price_per_share * Decimal(str(fully_diluted_shares))
    if pre_money <= 0:
        pre_money = Decimal('1000000.00')
    if price_per_share <= 0:
        price_per_share = Decimal('1.00')
    amount_raised = max((pre_money * Decimal('0.10')).quantize(Decimal('0.01')), Decimal('250000.00'))
    exit_values = defaults.get('default_exit_values') or ['25000000.00', '50000000.00', '100000000.00']

    return {
        'name': defaults.get('latest_round_name') or 'Board Scenario',
        'share_class': str(share_class.id),
        'investor_name': 'Board Scenario Investor',
        'pre_money_valuation': str(pre_money),
        'amount_raised': str(amount_raised),
        'price_per_share': str(price_per_share),
        'option_pool_top_up': 0,
        'apply_pro_rata': True,
        'include_anti_dilution': True,
        'exit_values': [str(value) for value in exit_values],
    }


def _build_equity_section(entities):
    profiles = WorkspaceEquityProfile.objects.filter(workspace__in=entities, equity_enabled=True).select_related('workspace')
    cards = []
    for profile in profiles:
        workspace = profile.workspace
        overview = get_scenario_overview(workspace)
        latest_report = workspace.equity_reports.filter(report_type='scenario_model').order_by('-created_at').first()
        latest_approval = workspace.equity_scenario_approvals.exclude(analysis_payload={}).order_by('-created_at').first()
        waterfall_source = {}
        fallback_generated = False
        fallback_input = None
        if latest_report:
            waterfall_source = latest_report.payload or {}
        elif latest_approval:
            waterfall_source = latest_approval.analysis_payload or {}
        analysis = waterfall_source.get('analysis', waterfall_source)
        waterfalls = analysis.get('waterfalls') or []
        if not waterfalls:
            fallback_input = _default_equity_scenario_input(workspace, overview)
            if fallback_input:
                try:
                    analysis = simulate_financing_scenario(workspace, fallback_input)
                    waterfalls = analysis.get('waterfalls') or []
                    fallback_generated = bool(waterfalls)
                except Exception:
                    waterfalls = []
        cards.append({
            'entity_id': workspace.id,
            'entity_name': workspace.name,
            'status': 'ready' if waterfalls else 'needs_scenario',
            'latest_report_title': latest_report.title if latest_report else '',
            'cap_table': overview.get('cap_table', {}),
            'defaults': overview.get('defaults', {}),
            'waterfalls': waterfalls[:3],
            'fallback_generated': fallback_generated,
            'fallback_scenario': fallback_input if fallback_generated else None,
        })
    return {
        'enabled_entities': profiles.count(),
        'entities': cards,
    }


def build_enterprise_reporting_dashboard(organization, entities, start, end):
    entity_summaries = [_entity_statement_summary(entity, start, end) for entity in entities]
    latest_consolidation = (
        Consolidation.objects.filter(organization=organization, status='completed')
        .prefetch_related('entities', 'intercompany_eliminations')
        .order_by('-consolidation_date', '-updated_at')
        .first()
    )

    consolidated_revenue = sum((row['revenue'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_expenses = sum((row['expenses'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_net_income = sum((row['net_income'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_cash = sum((row['cash'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_assets = sum((row['total_assets'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_liabilities = sum((row['liabilities'] for row in entity_summaries), DECIMAL_ZERO)
    consolidated_equity = sum((row['equity'] for row in entity_summaries), DECIMAL_ZERO)

    cashflow_months = _build_cashflow_months(entities, start, end)
    budget_section = _build_budget_section(entities, end)
    compliance_section = _build_compliance_section(organization, entities, start, end)
    scenario_section = _build_scenario_section(organization)
    equity_section = _build_equity_section(entities)

    return {
        'organization': {
            'id': organization.id,
            'name': organization.name,
            'currency': organization.primary_currency,
        },
        'period': {
            'start_date': start.isoformat(),
            'end_date': end.isoformat(),
        },
        'summary': {
            'entities_covered': len(entities),
            'revenue': _serialize_money(consolidated_revenue),
            'expenses': _serialize_money(consolidated_expenses),
            'net_income': _serialize_money(consolidated_net_income),
            'cash_on_hand': _serialize_money(consolidated_cash),
            'total_assets': _serialize_money(consolidated_assets),
            'total_liabilities': _serialize_money(consolidated_liabilities),
            'equity': _serialize_money(consolidated_equity),
        },
        'consolidated_statements': {
            'source': 'saved_consolidation' if latest_consolidation else 'live_aggregation',
            'consolidation_name': latest_consolidation.name if latest_consolidation else '',
            'consolidation_date': latest_consolidation.consolidation_date.isoformat() if latest_consolidation else '',
            'intercompany_eliminations': latest_consolidation.intercompany_eliminations.count() if latest_consolidation else 0,
            'balance_sheet': latest_consolidation.consolidated_balance_sheet if latest_consolidation else {
                'total_assets': _serialize_money(consolidated_assets),
                'total_liabilities': _serialize_money(consolidated_liabilities),
                'shareholders_equity': _serialize_money(consolidated_equity),
            },
            'profit_and_loss': latest_consolidation.consolidated_pnl if latest_consolidation else {
                'total_revenue': _serialize_money(consolidated_revenue),
                'total_expenses': _serialize_money(consolidated_expenses),
                'net_income': _serialize_money(consolidated_net_income),
            },
            'cashflow': latest_consolidation.consolidated_cashflow if latest_consolidation else {
                'net_cashflow': _serialize_money(sum((Decimal(str(month['forecast_net'])) for month in cashflow_months), DECIMAL_ZERO)),
            },
            'entities': [
                {
                    'entity_id': row['entity_id'],
                    'entity_name': row['entity_name'],
                    'revenue': _serialize_money(row['revenue']),
                    'expenses': _serialize_money(row['expenses']),
                    'net_income': _serialize_money(row['net_income']),
                    'cash': _serialize_money(row['cash']),
                    'liabilities': _serialize_money(row['liabilities']),
                }
                for row in entity_summaries
            ],
        },
        'cashflow_automation': {
            'forecast_horizon_months': len(cashflow_months),
            'timeline': cashflow_months,
            'automation': {
                'cashflow_forecast_rows': CashflowForecast.objects.filter(entity__in=entities).count(),
                'recurring_transaction_templates': RecurringTransaction.objects.filter(entity__in=entities, is_active=True).count(),
                'recurring_journal_templates': RecurringJournalTemplate.objects.filter(entity__in=entities, is_active=True).count(),
                'automation_workflows': AutomationWorkflow.objects.filter(organization=organization, is_active=True).count(),
            },
        },
        'budgeting_and_forecasting': budget_section,
        'variance_analysis': {
            'budget_variance': budget_section['summary'],
            'top_variances': budget_section['top_variances'],
        },
        'scenario_dashboard': scenario_section,
        'equity_waterfalls': equity_section,
        'automated_compliance_reports': compliance_section,
    }


def _audit_export(organization, *, entity=None, initiated_by=None, export_format='json', period=None, action='export'):
    AuditLog.objects.create(
        organization=organization,
        entity=entity,
        user=initiated_by,
        action=action,
        model_name='EnterpriseReportingPack',
        object_id=str(organization.id),
        changes={
            'format': export_format,
            'period': period or {},
        },
    )


def render_enterprise_reporting_pdf(payload):
    styles = getSampleStyleSheet()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=36, bottomMargin=24, leftMargin=32, rightMargin=32)
    story = []

    story.append(Paragraph(f"{payload['organization']['name']} Board Pack", styles['Title']))
    story.append(Paragraph(f"Reporting period: {payload['period']['start_date']} to {payload['period']['end_date']}", styles['Normal']))
    story.append(Spacer(1, 12))

    summary = payload.get('summary', {})
    summary_rows = [
        ['Metric', 'Value'],
        ['Revenue', str(summary.get('revenue', 0))],
        ['Expenses', str(summary.get('expenses', 0))],
        ['Net income', str(summary.get('net_income', 0))],
        ['Cash on hand', str(summary.get('cash_on_hand', 0))],
        ['Total assets', str(summary.get('total_assets', 0))],
        ['Total liabilities', str(summary.get('total_liabilities', 0))],
        ['Equity', str(summary.get('equity', 0))],
    ]
    summary_table = Table(summary_rows, hAlign='LEFT')
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17324d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    def add_section(title, rows):
        story.append(Paragraph(title, styles['Heading2']))
        if not rows:
            story.append(Paragraph('No data available for this section.', styles['Normal']))
            story.append(Spacer(1, 10))
            return
        table = Table(rows, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

    add_section(
        'Cashflow Timeline',
        [['Month', 'Actual Net', 'Forecast Net', 'Projected Close']] + [
            [row['month'], str(row['actual_net']), str(row['forecast_net']), str(row['projected_closing_cash'])]
            for row in payload.get('cashflow_automation', {}).get('timeline', [])[:12]
        ],
    )
    add_section(
        'Budget Variances',
        [['Entity', 'Category', 'Forecast', 'Variance']] + [
            [row['entity_name'], row['category'], str(row['forecast']), str(row['variance'])]
            for row in payload.get('variance_analysis', {}).get('top_variances', [])[:10]
        ],
    )
    add_section(
        'Scenario Dashboard',
        [['Scenario', 'Model', 'Type', 'Enterprise Value', 'IRR']] + [
            [row['name'], row['financial_model_name'], row['scenario_type'], str(row['enterprise_value']), str(row['irr'])]
            for row in payload.get('scenario_dashboard', {}).get('top_scenarios', [])[:10]
        ],
    )
    add_section(
        'Equity Waterfalls',
        [['Entity', 'Status', 'Waterfalls', 'Fallback']] + [
            [row['entity_name'], row['status'], str(len(row.get('waterfalls') or [])), 'Yes' if row.get('fallback_generated') else 'No']
            for row in payload.get('equity_waterfalls', {}).get('entities', [])
        ],
    )
    add_section(
        'Compliance Deadlines',
        [['Entity', 'Title', 'Type', 'Status', 'Due']] + [
            [row['entity_name'], row['title'], row['deadline_type'], row['status'], row['deadline_date']]
            for row in payload.get('automated_compliance_reports', {}).get('upcoming_deadlines', [])[:10]
        ],
    )

    doc.build(story)
    return buffer.getvalue()


def render_enterprise_reporting_xlsx(payload):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    summary_sheet['A1'] = f"{payload['organization']['name']} Enterprise Reporting Pack"
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet['A2'] = 'Period start'
    summary_sheet['B2'] = payload['period']['start_date']
    summary_sheet['A3'] = 'Period end'
    summary_sheet['B3'] = payload['period']['end_date']
    for row_index, (label, value) in enumerate(payload.get('summary', {}).items(), start=5):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)

    statements_sheet = workbook.create_sheet('Statements')
    statements_sheet.append(['Entity', 'Revenue', 'Expenses', 'Net Income', 'Cash', 'Liabilities'])
    for row in payload.get('consolidated_statements', {}).get('entities', []):
        statements_sheet.append([
            row['entity_name'], row['revenue'], row['expenses'], row['net_income'], row['cash'], row['liabilities'],
        ])

    cashflow_sheet = workbook.create_sheet('Cashflow')
    cashflow_sheet.append(['Month', 'Actual Net', 'Forecast Net', 'Opening Cash', 'Projected Closing'])
    for row in payload.get('cashflow_automation', {}).get('timeline', []):
        cashflow_sheet.append([
            row['month'], row['actual_net'], row['forecast_net'], row['opening_cash'], row['projected_closing_cash'],
        ])

    budget_sheet = workbook.create_sheet('Budget Variance')
    budget_sheet.append(['Entity', 'Category', 'Limit', 'Spent', 'Forecast', 'Variance'])
    for row in payload.get('budgeting_and_forecasting', {}).get('categories', []):
        budget_sheet.append([
            row['entity_name'], row['category'], row['limit'], row['spent'], row['forecast'], row['variance'],
        ])

    scenario_sheet = workbook.create_sheet('Scenarios')
    scenario_sheet.append(['Scenario', 'Model', 'Type', 'Enterprise Value', 'IRR', 'Probability'])
    for row in payload.get('scenario_dashboard', {}).get('top_scenarios', []):
        scenario_sheet.append([
            row['name'], row['financial_model_name'], row['scenario_type'], row['enterprise_value'], row['irr'], row['probability'],
        ])

    equity_sheet = workbook.create_sheet('Equity')
    equity_sheet.append(['Entity', 'Status', 'Fallback Generated', 'Latest Report', 'Waterfall Count'])
    for row in payload.get('equity_waterfalls', {}).get('entities', []):
        equity_sheet.append([
            row['entity_name'], row['status'], row.get('fallback_generated', False), row.get('latest_report_title', ''), len(row.get('waterfalls') or []),
        ])

    compliance_sheet = workbook.create_sheet('Compliance')
    compliance_sheet.append(['Entity', 'Title', 'Type', 'Status', 'Due'])
    for row in payload.get('automated_compliance_reports', {}).get('upcoming_deadlines', []):
        compliance_sheet.append([
            row['entity_name'], row['title'], row['deadline_type'], row['status'], row['deadline_date'],
        ])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_enterprise_reporting_payload(payload, export_format):
    export_format = (export_format or 'json').lower()
    org_slug = payload['organization']['name'].lower().replace(' ', '-')
    period_end = payload['period']['end_date']
    if export_format == 'pdf':
        return render_enterprise_reporting_pdf(payload), 'application/pdf', f'{org_slug}-board-pack-{period_end}.pdf'
    if export_format == 'xlsx':
        return render_enterprise_reporting_xlsx(payload), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'{org_slug}-board-pack-{period_end}.xlsx'
    return (
        BytesIO(json.dumps(payload, indent=2).encode('utf-8')).getvalue(),
        'application/json',
        f'{org_slug}-board-pack-{period_end}.json',
    )


def _parse_datetime(value):
    if not value:
        return None
    try:
        parsed = timezone.datetime.fromisoformat(value.replace('Z', '+00:00'))
        if timezone.is_naive(parsed):
            parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
        return parsed
    except Exception:
        return None


def _resolve_schedule_timezone_name(trigger_config):
    timezone_name = (trigger_config or {}).get('schedule_timezone') or DEFAULT_SCHEDULE_TIMEZONE
    try:
        ZoneInfo(timezone_name)
        return timezone_name
    except Exception:
        return DEFAULT_SCHEDULE_TIMEZONE


def _resolve_schedule_timezone(trigger_config):
    return ZoneInfo(_resolve_schedule_timezone_name(trigger_config))


def _resolve_retention_days(trigger_config):
    try:
        retention_days = int((trigger_config or {}).get('retention_days') or DEFAULT_ARTIFACT_RETENTION_DAYS)
    except (TypeError, ValueError):
        retention_days = DEFAULT_ARTIFACT_RETENTION_DAYS
    return max(1, retention_days)


def normalize_schedule_trigger_config(trigger_config, *, now=None):
    now = now or timezone.now()
    normalized = dict(trigger_config or {})
    if normalized.get('frequency') not in {'daily', 'weekly', 'monthly', 'quarterly'}:
        normalized['frequency'] = 'monthly'
    normalized['schedule_timezone'] = _resolve_schedule_timezone_name(normalized)
    normalized['retention_days'] = _resolve_retention_days(normalized)
    normalized['next_run_at'] = (_parse_datetime(normalized.get('next_run_at')) or now).isoformat()
    return normalized


def _resolve_report_period(action, now):
    end = _parse_datetime(action.get('end_date'))
    end_date = end.date() if end else now.date()
    start = _parse_datetime(action.get('start_date'))
    if start:
        start_date = start.date()
    else:
        months_back = int(action.get('months_back') or 12)
        start_date = end_date - timedelta(days=max(1, months_back) * 30)
    return start_date, end_date


def _next_run_at(trigger_config, *, now=None):
    now = now or timezone.now()
    frequency = (trigger_config or {}).get('frequency', 'monthly')
    current = _parse_datetime((trigger_config or {}).get('next_run_at')) or now
    local_timezone = _resolve_schedule_timezone(trigger_config)
    local_current = current.astimezone(local_timezone)

    if frequency == 'daily':
        local_next = local_current + relativedelta(days=1)
    elif frequency == 'weekly':
        local_next = local_current + relativedelta(weeks=1)
    elif frequency == 'quarterly':
        local_next = local_current + relativedelta(months=3)
    else:
        local_next = local_current + relativedelta(months=1)

    return local_next.astimezone(datetime_timezone.utc)


def _due_for_execution(workflow, *, now=None):
    now = now or timezone.now()
    if not workflow.is_active or workflow.trigger_type != 'schedule':
        return False
    next_run_at = _parse_datetime((workflow.trigger_config or {}).get('next_run_at'))
    if next_run_at is None:
        return True
    return next_run_at <= now


def execute_automation_workflow(workflow, *, initiated_by=None, trigger_type='manual'):
    now = timezone.now()
    execution = AutomationExecution.objects.create(workflow=workflow, status='running', started_at=now)
    log_platform_audit_event(
        domain='automation',
        actor=initiated_by,
        organization=workflow.organization,
        entity=workflow.entity,
        event_type='automation_workflow.started',
        action='workflow_run_started',
        resource_type='AutomationWorkflow',
        resource_id=str(workflow.id),
        subject_type='automation_workflow',
        subject_id=str(workflow.id),
        resource_name=workflow.name,
        summary=f'Automation workflow started: {workflow.name}',
        context={'trigger_type': trigger_type, 'execution_id': execution.id},
    )
    try:
        entities = [workflow.entity] if workflow.entity_id else list(workflow.organization.entities.filter(status='active'))
        action_results = []
        for action in workflow.actions or []:
            action_type = action.get('type')
            if action_type not in SUPPORTED_REPORT_ACTIONS:
                action_results.append({'type': action_type, 'status': 'skipped', 'reason': 'unsupported_action'})
                continue

            start_date, end_date = _resolve_report_period(action, now)
            payload = build_enterprise_reporting_dashboard(workflow.organization, entities, start_date, end_date)
            export_format = action.get('format', 'pdf')
            content, content_type, filename = export_enterprise_reporting_payload(payload, export_format)
            artifact_type = 'compliance_board_pack' if action_type == 'compliance_reporting_pack' else 'enterprise_board_pack'
            artifact = AutomationArtifact(
                workflow=workflow,
                execution=execution,
                organization=workflow.organization,
                entity=workflow.entity,
                artifact_type=artifact_type,
                export_format=export_format,
                file_name=filename,
                metadata={
                    'period': payload['period'],
                    'summary': payload.get('summary', {}),
                    'action_type': action_type,
                },
                generated_by=initiated_by,
            )
            artifact.file_path.save(filename, ContentFile(content), save=False)
            artifact.save()
            recipients = action.get('recipients') or ([workflow.organization.owner.email] if workflow.organization.owner and workflow.organization.owner.email else [])
            delivered = False
            if recipients:
                email = EmailMessage(
                    subject=action.get('subject') or f"{workflow.organization.name} Enterprise Reporting Pack",
                    body=action.get('body') or f"Attached is the {export_format.upper()} enterprise reporting pack for {workflow.organization.name} covering {payload['period']['start_date']} to {payload['period']['end_date']}.",
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@atonixcorp.local'),
                    to=recipients,
                )
                email.attach(filename, content, content_type)
                email.send(fail_silently=False)
                delivered = True

            _audit_export(
                workflow.organization,
                entity=workflow.entity,
                initiated_by=initiated_by,
                export_format=export_format,
                period=payload['period'],
            )
            action_results.append({
                'type': action_type,
                'status': 'completed',
                'format': export_format,
                'filename': filename,
                'artifact_id': artifact.id,
                'recipient_count': len(recipients),
                'delivered': delivered,
                'period': payload['period'],
            })

        trigger_config = dict(workflow.trigger_config or {})
        if workflow.trigger_type == 'schedule':
            trigger_config['last_trigger_type'] = trigger_type
            trigger_config['last_run_at'] = now.isoformat()
            trigger_config['schedule_timezone'] = _resolve_schedule_timezone_name(trigger_config)
            trigger_config['retention_days'] = _resolve_retention_days(trigger_config)
            trigger_config['next_run_at'] = _next_run_at(trigger_config, now=now).isoformat()
            workflow.trigger_config = trigger_config
            workflow.save(update_fields=['trigger_config', 'updated_at'])

        execution.status = 'completed'
        execution.completed_at = timezone.now()
        execution.execution_result = {
            'workflow_id': workflow.id,
            'trigger_type': trigger_type,
            'actions': action_results,
        }
        execution.save(update_fields=['status', 'completed_at', 'execution_result'])
        log_platform_audit_event(
            domain='automation',
            actor=initiated_by,
            organization=workflow.organization,
            entity=workflow.entity,
            event_type='automation_workflow.completed',
            action='workflow_run_completed',
            resource_type='AutomationWorkflow',
            resource_id=str(workflow.id),
            subject_type='automation_workflow',
            subject_id=str(workflow.id),
            resource_name=workflow.name,
            summary=f'Automation workflow completed: {workflow.name}',
            context={'trigger_type': trigger_type, 'execution_id': execution.id},
            metadata={'actions': action_results},
        )
        return execution
    except Exception as exc:
        execution.status = 'failed'
        execution.completed_at = timezone.now()
        execution.error_message = str(exc)
        execution.save(update_fields=['status', 'completed_at', 'error_message'])
        log_platform_audit_event(
            domain='automation',
            actor=initiated_by,
            organization=workflow.organization,
            entity=workflow.entity,
            event_type='automation_workflow.failed',
            action='workflow_run_failed',
            resource_type='AutomationWorkflow',
            resource_id=str(workflow.id),
            subject_type='automation_workflow',
            subject_id=str(workflow.id),
            resource_name=workflow.name,
            summary=f'Automation workflow failed: {workflow.name}',
            context={'trigger_type': trigger_type, 'execution_id': execution.id},
            metadata={'error_message': str(exc)},
        )
        raise


def run_due_automation_workflows(*, now=None):
    now = now or timezone.now()
    completed = 0
    failed = 0
    skipped = 0
    for workflow in AutomationWorkflow.objects.filter(is_active=True, trigger_type='schedule').select_related('organization', 'entity', 'organization__owner'):
        if not _due_for_execution(workflow, now=now):
            skipped += 1
            continue
        try:
            execute_automation_workflow(workflow, initiated_by=None, trigger_type='scheduled')
            completed += 1
        except Exception:
            failed += 1
    cleanup = cleanup_automation_artifacts(now=now)
    return {
        'completed': completed,
        'failed': failed,
        'skipped': skipped,
        'artifacts_deleted': cleanup['deleted'],
        'bytes_reclaimed': cleanup['bytes_reclaimed'],
    }


def cleanup_automation_artifacts(*, now=None):
    now = now or timezone.now()
    deleted = 0
    bytes_reclaimed = 0

    queryset = AutomationArtifact.objects.select_related('workflow')
    for artifact in queryset.iterator():
        retention_days = _resolve_retention_days(getattr(artifact.workflow, 'trigger_config', {}) or {})
        cutoff = now - timedelta(days=retention_days)
        if artifact.created_at >= cutoff:
            continue

        storage = artifact.file_path.storage
        file_name = artifact.file_path.name
        if file_name and storage.exists(file_name):
            try:
                bytes_reclaimed += storage.size(file_name)
            except Exception:
                pass
            storage.delete(file_name)
        artifact.delete()
        deleted += 1

    return {
        'deleted': deleted,
        'bytes_reclaimed': bytes_reclaimed,
    }


def _get_artifact_size_bytes(artifact):
    file_name = getattr(artifact.file_path, 'name', '')
    if not file_name:
        return 0
    try:
        return artifact.file_path.storage.size(file_name)
    except Exception:
        return 0


def build_automation_cleanup_impact_report(*, workflows, now=None, days_ahead=30):
    now = now or timezone.now()
    try:
        days_ahead = int(days_ahead)
    except (TypeError, ValueError):
        days_ahead = 30
    days_ahead = max(1, days_ahead)
    horizon = now + timedelta(days=days_ahead)

    workflows = list(workflows)
    artifacts_by_workflow = defaultdict(list)
    for artifact in AutomationArtifact.objects.filter(workflow__in=workflows).select_related('workflow').iterator():
        artifacts_by_workflow[artifact.workflow_id].append(artifact)

    workflow_rows = []
    totals = {
        'workflow_count': len(workflows),
        'artifacts_total': 0,
        'artifacts_expiring': 0,
        'artifacts_expired': 0,
        'bytes_total': 0,
        'bytes_expiring': 0,
    }

    for workflow in workflows:
        retention_days = _resolve_retention_days(workflow.trigger_config)
        timezone_name = _resolve_schedule_timezone_name(workflow.trigger_config)
        workflow_artifacts = artifacts_by_workflow.get(workflow.id, [])

        total_artifacts = 0
        total_bytes = 0
        expiring_artifacts = 0
        expired_artifacts = 0
        expiring_bytes = 0
        next_expiration_at = None

        for artifact in workflow_artifacts:
            total_artifacts += 1
            size_bytes = _get_artifact_size_bytes(artifact)
            total_bytes += size_bytes
            expires_at = artifact.created_at + timedelta(days=retention_days)
            if next_expiration_at is None or expires_at < next_expiration_at:
                next_expiration_at = expires_at
            if expires_at <= horizon:
                expiring_artifacts += 1
                expiring_bytes += size_bytes
                if expires_at <= now:
                    expired_artifacts += 1

        totals['artifacts_total'] += total_artifacts
        totals['artifacts_expiring'] += expiring_artifacts
        totals['artifacts_expired'] += expired_artifacts
        totals['bytes_total'] += total_bytes
        totals['bytes_expiring'] += expiring_bytes

        workflow_rows.append({
            'workflow_id': workflow.id,
            'workflow_name': workflow.name,
            'entity_name': workflow.entity.name if workflow.entity_id else 'All organization entities',
            'retention_days': retention_days,
            'schedule_timezone': timezone_name,
            'total_artifacts': total_artifacts,
            'total_bytes': total_bytes,
            'artifacts_expiring_within_window': expiring_artifacts,
            'artifacts_already_expired': expired_artifacts,
            'bytes_expiring_within_window': expiring_bytes,
            'next_expiration_at': next_expiration_at.isoformat() if next_expiration_at else None,
        })

    workflow_rows.sort(
        key=lambda row: (
            -row['artifacts_already_expired'],
            -row['artifacts_expiring_within_window'],
            row['next_expiration_at'] or '9999-12-31T00:00:00+00:00',
            row['workflow_name'].lower(),
        )
    )

    return {
        'generated_at': now.isoformat(),
        'days_ahead': days_ahead,
        'summary': totals,
        'workflows': workflow_rows,
    }